import Config
import openpyxl
import csv
from resource_classes import Organisation, Biobank, Patientregistry, Dataset, Distribution


class TemplateReader:
    """
    """
    def get_datasets(self):
        """
        This method creates datasets objects by extracting content from the dataset input CSV file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/LUMC-BioSemantics/EJP-RD-WP19-FDP-template>

        :return: Dict of datasets
        """

        reader = csv.reader(open(Config.DATASET_INPUT_FILE, 'r'))
        catalog_url = Config.CATALOG_URL
        datasets = {}
        for row in reader:
            if reader.line_num > 1:
                print(row)
                title = row[0]
                publisher_url = row[1]
                description = row[2]
                language = row[3]
                license = row[4]
                contact_point = row[5]
                landing_page = row[6]
                keywords_str = row[7]
                themes_str = row[8]
                language_url = None
                license_url = None
                landing_page_url = None
                contact_point_url = None

                if not description:
                    description = "Metadata od dataset " + title
                # Create language triples
                if language:
                    language_url = "http://id.loc.gov/vocabulary/iso639-1/" + language.strip()
                # Create license triples
                if license:
                    license_url = license.strip()
                # Create license triples
                if landing_page:
                    landing_page_url = landing_page.strip()
                # Create contact point triples
                if contact_point:
                    contact_point_url = contact_point.strip()
                # Create keywords list
                keywords = []
                for keyword in keywords_str.split(","):
                    keyword = keyword.strip()
                    keywords.append(keyword)
                # Create themes list
                themes = []
                for theme in themes_str.split(","):
                    theme = theme.strip()
                    themes.append(theme)
                dataset = Dataset.Dataset(catalog_url, title, description, keywords, themes, publisher_url,
                                          language_url, license_url, landing_page_url, contact_point_url)
                datasets[title] = dataset
        return datasets

    def get_distributions(self):
        """
        This method creates distribution objects by extracting content from the distribution input CSV file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/LUMC-BioSemantics/EJP-RD-WP19-FDP-template>

        :return: Dict of distribution
        """

        reader = csv.reader(open(Config.DISTRIBUTION_INPUT_FILE, 'r'))
        distributions = {}
        for row in reader:
            if reader.line_num > 1:
                print(row)
                title = row[0]
                dataset_name = row[1]
                publisher_url = row[2]
                description = row[3]
                language = row[4]
                license = row[5]
                access_url = row[6]
                download_url = row[7]
                media_type = row[8]
                compression_format = row[9]
                format = row[10]
                byte_size = row[11]
                language_url = None
                license_url = None

                if not description:
                    description = "Metadata od dataset " + title
                # Create language triples
                if language:
                    language_url = "http://id.loc.gov/vocabulary/iso639-1/" + language.strip()
                # Create license triples
                if license:
                    license_url = license.strip()
                distribution = Distribution.Distribution(None, title, description, publisher_url, language_url,
                                                         license_url, access_url, download_url, media_type,
                                                         compression_format, format, byte_size, dataset_name)
                distributions[title] = distribution
        return distributions
    
    def get_organisations(self):
        """
        This method creates organisation objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of organisations
        """
        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['Organisation']
        
        # Loop over rows of excel sheet
        first_row = True
        organisations = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                continue

            if row[0].value != None:
                # Retrieve field values from excel files
                title = row[0].value
                description = row[1].value

                if type(row[2].value) == str:
                    pages = [page.strip() for page in row[2].value.split(";")]
                else:
                    pages = []

                location_title = row[3].value
                location_description = row[4].value

                # Create organisation object and add to organisation dictionary
                organisation = Organisation.Organisation(Config.CATALOG_URL, title, description, location_title, location_description, pages)
                organisations[organisation.TITLE] = organisation

        return organisations

    def get_biobanks(self):
        """
        This method creates biobank objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of biobanks
        """
        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['BiobankPatientRegistry']
        
        # Loop over rows of excel sheet
        first_row = True
        biobanks = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                continue

            if row[0].value != None:
                # Retrieve field values from excel files
                title = row[0].value
                description = row[1].value
                populationcoverage = row[2].value

                if type(row[3].value) == str:
                    themes = [theme.strip() for theme in row[3].value.split(";")]
                else:
                    themes = []

                publisher_name = row[4].value

                if type(row[5].value) == str:
                    pages = [page.strip() for page in row[5].value.split(";")]
                else:
                    pages = []

                resource_type = row[6].value

                # Create biobank object and add to biobank dictionary if it is a biobank
                if resource_type == "Biobank":
                    biobank = Biobank.Biobank(Config.CATALOG_URL, None, title, description, populationcoverage, themes, publisher_name, pages)
                    biobanks[biobank.TITLE] = biobank

        return biobanks

    def get_patientregistries(self):
        """
        This method creates patient registry objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of patientregistries
        """
        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['BiobankPatientRegistry']
        
        # Loop over rows of excel sheet
        first_row = True
        patientregistries = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                continue

            if row[0].value != None:
                # Retrieve field values from excel files
                title = row[0].value
                description = row[1].value
                populationcoverage = row[2].value

                if type(row[3].value) == str:
                    themes = [theme.strip() for theme in row[3].value.split(";")]
                else:
                    themes = []

                publisher_name = row[4].value

                if type(row[5].value) == str:
                    pages = [page.strip() for page in row[5].value.split(";")]
                else:
                    pages = []

                resource_type = row[6].value

                # Create patient registry object and add to patientregistry dictionary if it is a patientregistry
                if resource_type == "Patient registry":
                    patientregistry = Patientregistry.Patientregistry(Config.CATALOG_URL, None, title, description, populationcoverage, themes, publisher_name, pages)
                    patientregistries[patientregistry.TITLE] = patientregistry

        return patientregistries

    def get_datasets_excel(self):
        """
        This method creates dataset objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of datasets
        """
        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['BiobankPatientRegistry']
        
        # Loop over rows of excel sheet
        first_row = True
        datasets = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                continue

            if row[0].value != None:
                # Retrieve field values from excel files
                title = row[0].value
                description = row[1].value

                if type(row[2].value) == str:
                    themes = [theme.strip() for theme in row[2].value.split(";")]
                else:
                    themes = []

                vpconnection = row[3].value
                license = row[4].value
                
                if type(row[5].value) == str:
                    related = [item.strip() for item in row[5].value.split(";")]
                else:
                    related = []

                version = row[6].value

                if type(row[7].value) == str:
                    keywords = [item.strip() for item in row[7].value.split(";")]
                else:
                    keywords = []

                publisher = row[8].value
                page = row[9].value

                # Create dataset object and add to dataset dictionary
                dataset = Dataset.Dataset(Config.CATALOG_URL, title, description, keywords, themes, publisher, "en", license, page, None)
                datasets[dataset.TITLE] = dataset

        return datasets